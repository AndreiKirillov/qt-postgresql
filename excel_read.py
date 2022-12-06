import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import psycopg2

connection = psycopg2.connect(
  database="oil_gas_field", 
  user="myuser", 
  password="123", 
  host="127.0.0.1"
)
cursor = connection.cursor()

# Выполнение SQL-запроса
cursor.execute("SELECT version();")
# Получить результат
record = cursor.fetchone()
print("Вы подключены к - ", record, "\n")

# загружаем содержимое excel таблицы
wb = load_workbook('codes.xlsx')

# Get a sheet by name 
sheet = wb.get_sheet_by_name('Лист1')
max_row = sheet.max_row + 1

now = datetime.now()
date_time = now.strftime("%Y-%m-%d %H:%M:%S")

# Print out values in column 2 
for i in range(2, max_row):
    value_type = sheet.cell(row=i, column=5).value
    cipher = int(sheet.cell(row=i, column=6).value)
    value_str = sheet.cell(row=i, column=7).value
    if value_type == 'System.Int32':
    	value_int = int(value_str)
    	insert_query = f"INSERT INTO tech_object_data (reading_time, technology_object_id, sensor_id, int_value) VALUES (\'{date_time}\', 1, {cipher}, {value_int})"
    	print(insert_query)
    	cursor.execute(insert_query)
    	connection.commit()
    else:
    	value_double = float(value_str)
    	insert_query = f"INSERT INTO tech_object_data (reading_time, technology_object_id, sensor_id, double_value) VALUES (\'{date_time}\', 1, {cipher}, {value_double})"
    	print(insert_query)
    	cursor.execute(insert_query)
    	connection.commit()

record = cursor.fetchall()

cursor.close()
connection.close()


